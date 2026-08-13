from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from .normalization import normalize_company_name
from ..jobs_repository import _conn

EMPLOYER_HEADERS = ("EMPLOYER_NAME", "EMPLOYER_BUSINESS_NAME")
STATUS_HEADERS = ("CASE_STATUS", "STATUS")
VISA_HEADERS = ("VISA_CLASS", "VISA_TYPE")

def _norm_header(value):
    return re.sub(r"[^A-Z0-9]+", "_", str(value or "").upper()).strip("_")

def _find_header(headers, candidates):
    normalized = {_norm_header(h): i for i, h in enumerate(headers)}
    for c in candidates:
        if c in normalized:
            return normalized[c]
    return None

def _strength(total, recent):
    if recent >= 20 or total >= 100:
        return "STRONG"
    if recent >= 5 or total >= 25:
        return "MEDIUM"
    if recent >= 1 or total >= 5:
        return "LOW"
    return "UNKNOWN"

def _init_rollup():
    with _conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sponsor_rollup (
                normalized_name TEXT PRIMARY KEY,
                display_name TEXT NOT NULL,
                source TEXT NOT NULL DEFAULT 'DOL_OFLC_LCA',
                total_filings INTEGER NOT NULL DEFAULT 0,
                approved_count INTEGER NOT NULL DEFAULT 0,
                denied_count INTEGER NOT NULL DEFAULT 0,
                recent_filings INTEGER NOT NULL DEFAULT 0,
                latest_year INTEGER,
                sponsor_strength TEXT NOT NULL DEFAULT 'UNKNOWN',
                last_verified_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()

def _aggregate(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)

    employer_idx = _find_header(headers, EMPLOYER_HEADERS)
    status_idx = _find_header(headers, STATUS_HEADERS)
    visa_idx = _find_header(headers, VISA_HEADERS)

    if employer_idx is None:
        raise RuntimeError(f"Could not find employer column in {path.name}")

    out = defaultdict(lambda: {"display_name":"", "filings":0, "approved":0, "denied":0})

    for row in rows:
        if employer_idx >= len(row):
            continue
        raw_name = str(row[employer_idx] or "").strip()
        if not raw_name:
            continue

        if visa_idx is not None and visa_idx < len(row):
            visa = str(row[visa_idx] or "").upper().strip()
            if visa and visa not in {"H-1B", "H1B"}:
                continue

        key = normalize_company_name(raw_name)
        if not key:
            continue

        item = out[key]
        item["display_name"] = item["display_name"] or raw_name
        item["filings"] += 1

        status = ""
        if status_idx is not None and status_idx < len(row):
            status = str(row[status_idx] or "").upper()

        if "CERTIFIED" in status or "APPROVED" in status:
            item["approved"] += 1
        elif "DENIED" in status:
            item["denied"] += 1

    wb.close()
    return dict(out)

def import_local_files(fy2026_q2: str, fy2025_q4: str):
    _init_rollup()

    files = {
        2026: Path(fy2026_q2).expanduser().resolve(),
        2025: Path(fy2025_q4).expanduser().resolve(),
    }

    for year, path in files.items():
        if not path.exists():
            raise FileNotFoundError(f"FY{year} file not found: {path}")

    combined = defaultdict(lambda: {
        "display_name":"", "total":0, "approved":0, "denied":0, "recent":0
    })

    file_stats = []

    for year, path in files.items():
        data = _aggregate(path)
        file_stats.append({"year":year, "file":path.name, "employers":len(data)})

        for key, item in data.items():
            out = combined[key]
            out["display_name"] = out["display_name"] or item["display_name"]
            out["total"] += item["filings"]
            out["approved"] += item["approved"]
            out["denied"] += item["denied"]
            if year == 2026:
                out["recent"] += item["filings"]

    with _conn() as conn:
        for key, item in combined.items():
            conn.execute("""
                INSERT INTO sponsor_rollup (
                    normalized_name, display_name, source,
                    total_filings, approved_count, denied_count,
                    recent_filings, latest_year, sponsor_strength, last_verified_at
                )
                VALUES (?, ?, 'DOL_OFLC_LCA', ?, ?, ?, ?, 2026, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(normalized_name) DO UPDATE SET
                    display_name=excluded.display_name,
                    total_filings=excluded.total_filings,
                    approved_count=excluded.approved_count,
                    denied_count=excluded.denied_count,
                    recent_filings=excluded.recent_filings,
                    latest_year=excluded.latest_year,
                    sponsor_strength=excluded.sponsor_strength,
                    last_verified_at=CURRENT_TIMESTAMP
            """, (
                key,
                item["display_name"],
                item["total"],
                item["approved"],
                item["denied"],
                item["recent"],
                _strength(item["total"], item["recent"]),
            ))
        conn.commit()

    return {
        "files_processed": file_stats,
        "sponsor_rollup_rows": len(combined),
    }
