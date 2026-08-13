from __future__ import annotations

import re
import tempfile
from collections import defaultdict
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from .normalization import normalize_company_name
from ..jobs_repository import _conn

DOL_PERFORMANCE_URL = "https://www.dol.gov/agencies/eta/foreign-labor/performance"

TARGET_FILES = {
    2026: ("FY2026", "Q2"),
    2025: ("FY2025", "Q4"),
}

EMPLOYER_HEADERS = ("EMPLOYER_NAME", "EMPLOYER_BUSINESS_NAME")
STATUS_HEADERS = ("CASE_STATUS", "STATUS")
VISA_HEADERS = ("VISA_CLASS", "VISA_TYPE")

def _norm_header(value):
    return re.sub(r"[^A-Z0-9]+", "_", str(value or "").upper()).strip("_")

def _find_header(headers, candidates):
    normalized = {_norm_header(h): i for i, h in enumerate(headers)}
    for candidate in candidates:
        if candidate in normalized:
            return normalized[candidate]
    return None

def init_rollup():
    with _conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sponsor_rollup (
                normalized_name TEXT PRIMARY KEY,
                display_name TEXT NOT NULL,
                source TEXT NOT NULL DEFAULT 'DOL_OFLC_LCA',
                total_filings INTEGER NOT NULL DEFAULT 0,
                approved_count INTEGER NOT NULL DEFAULT 0,
                denied_count INTEGER NOT NULL DEFAULT 0,
                recent_filings INTEGER NOT NULL DEFAULT 0,
                latest_year INTEGER,
                sponsor_strength TEXT NOT NULL DEFAULT 'UNKNOWN',
                last_verified_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()

def _strength(total, recent):
    if recent >= 20 or total >= 100:
        return "STRONG"
    if recent >= 5 or total >= 25:
        return "MEDIUM"
    if recent >= 1 or total >= 5:
        return "LOW"
    return "UNKNOWN"

def discover_dol_files():
    r = requests.get(DOL_PERFORMANCE_URL, timeout=30, headers={"User-Agent":"visa-job-finder/1.0"})
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")

    links = []
    for a in soup.find_all("a", href=True):
        href = urljoin(DOL_PERFORMANCE_URL, a["href"])
        label = " ".join(a.stripped_strings)
        combined = f"{label} {href}"
        upper = combined.upper()
        if "LCA" in upper and ".XLSX" in upper:
            links.append((upper, href))

    found = {}
    for year, (fy, quarter) in TARGET_FILES.items():
        for label, href in links:
            if fy in label and quarter in label and "APPENDIX" not in label and "WORKSITE" not in label:
                if "DISCLOSURE" in label or "DISLCLOSURE" in label:
                    found[year] = href
                    break
    return found

def _download(url):
    r = requests.get(url, timeout=120, headers={"User-Agent":"visa-job-finder/1.0"})
    r.raise_for_status()
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(r.content)
    tmp.close()
    return Path(tmp.name)

def aggregate_xlsx(path, year):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)

    employer_idx = _find_header(headers, EMPLOYER_HEADERS)
    status_idx = _find_header(headers, STATUS_HEADERS)
    visa_idx = _find_header(headers, VISA_HEADERS)

    if employer_idx is None:
        raise RuntimeError(f"Could not find employer column in FY{year} LCA file")

    agg = defaultdict(lambda: {"display_name":"", "filings":0, "approved":0, "denied":0})

    for row in rows:
        raw_name = str(row[employer_idx] or "").strip() if employer_idx < len(row) else ""
        if not raw_name:
            continue

        if visa_idx is not None and visa_idx < len(row):
            visa = str(row[visa_idx] or "").upper().strip()
            if visa and visa not in {"H-1B", "H1B"}:
                continue

        key = normalize_company_name(raw_name)
        if not key:
            continue

        item = agg[key]
        item["display_name"] = item["display_name"] or raw_name
        item["filings"] += 1

        status = str(row[status_idx] or "").upper() if status_idx is not None and status_idx < len(row) else ""
        if "CERTIFIED" in status or "APPROVED" in status:
            item["approved"] += 1
        elif "DENIED" in status:
            item["denied"] += 1

    wb.close()
    return dict(agg)

def import_dol_lca():
    init_rollup()
    urls = discover_dol_files()
    if not urls:
        raise RuntimeError("Could not discover official DOL LCA disclosure files.")

    combined = defaultdict(lambda: {
        "display_name":"", "total":0, "approved":0, "denied":0,
        "recent":0, "latest_year":None
    })

    processed = []
    latest_year = max(urls)

    for year in sorted(urls):
        path = _download(urls[year])
        try:
            data = aggregate_xlsx(path, year)
        finally:
            path.unlink(missing_ok=True)

        processed.append({"year":year, "employers":len(data)})

        for key, item in data.items():
            out = combined[key]
            out["display_name"] = out["display_name"] or item["display_name"]
            out["total"] += item["filings"]
            out["approved"] += item["approved"]
            out["denied"] += item["denied"]
            out["latest_year"] = max(out["latest_year"] or year, year)
            if year == latest_year:
                out["recent"] += item["filings"]

    with _conn() as conn:
        for key, item in combined.items():
            conn.execute("""
                INSERT INTO sponsor_rollup (
                    normalized_name, display_name, source,
                    total_filings, approved_count, denied_count,
                    recent_filings, latest_year, sponsor_strength, last_verified_at
                )
                VALUES (?, ?, 'DOL_OFLC_LCA', ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(normalized_name) DO UPDATE SET
                    display_name=excluded.display_name,
                    total_filings=excluded.total_filings,
                    approved_count=excluded.approved_count,
                    denied_count=excluded.denied_count,
                    recent_filings=excluded.recent_filings,
                    latest_year=excluded.latest_year,
                    sponsor_strength=excluded.sponsor_strength,
                    last_verified_at=CURRENT_TIMESTAMP
            """, (
                key, item["display_name"], item["total"], item["approved"],
                item["denied"], item["recent"], item["latest_year"],
                _strength(item["total"], item["recent"])
            ))
        conn.commit()

    return {
        "source":"U.S. Department of Labor OFLC LCA disclosure data",
        "files_processed":processed,
        "sponsor_rollup_rows":len(combined)
    }
